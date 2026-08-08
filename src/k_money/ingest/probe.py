"""추출된 파일의 구조를 들여다본다.

뱅크샐러드 내보내기의 실제 스키마를 모르는 상태이므로, 인제스터를 쓰기 전에
시트 이름 · 헤더 · 샘플 행을 먼저 찍어보고 스펙을 확정한다.
"""

from __future__ import annotations

import csv
import json
from dataclasses import asdict, dataclass, field
from pathlib import Path

from openpyxl import load_workbook


@dataclass
class SheetShape:
    name: str
    max_row: int
    max_column: int
    header: list[str] = field(default_factory=list)
    sample_rows: list[list[str]] = field(default_factory=list)


@dataclass
class FileShape:
    path: str
    kind: str  # "xlsx" | "csv" | "unknown"
    sheets: list[SheetShape] = field(default_factory=list)
    note: str = ""


def _cell(value) -> str:
    if value is None:
        return ""
    return str(value)


def probe_xlsx(path: Path, sample: int = 3) -> FileShape:
    wb = load_workbook(path, read_only=True, data_only=True)
    shape = FileShape(path=str(path), kind="xlsx")
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            header: list[str] = []
            samples: list[list[str]] = []
            for idx, row in enumerate(rows):
                cells = [_cell(c) for c in row]
                if idx == 0:
                    header = cells
                elif len(samples) < sample:
                    samples.append(cells)
                else:
                    break
            shape.sheets.append(
                SheetShape(
                    name=ws.title,
                    max_row=ws.max_row or 0,
                    max_column=ws.max_column or 0,
                    header=header,
                    sample_rows=samples,
                )
            )
    finally:
        wb.close()
    return shape


def probe_csv(path: Path, sample: int = 3) -> FileShape:
    shape = FileShape(path=str(path), kind="csv")
    # 뱅샐이 UTF-8 BOM 이나 CP949 로 뽑을 수 있어 순서대로 시도한다.
    for encoding in ("utf-8-sig", "utf-8", "cp949"):
        try:
            with path.open(newline="", encoding=encoding) as fh:
                reader = csv.reader(fh)
                header = next(reader, [])
                samples = [row for _, row in zip(range(sample), reader)]
                total = sum(1 for _ in reader) + len(samples) + 1
            shape.sheets.append(
                SheetShape(
                    name=f"(csv:{encoding})",
                    max_row=total,
                    max_column=len(header),
                    header=header,
                    sample_rows=samples,
                )
            )
            return shape
        except UnicodeDecodeError:
            continue
    shape.note = "알려진 인코딩(utf-8/utf-8-sig/cp949)으로 읽지 못했습니다."
    return shape


def probe(path: Path) -> FileShape:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xlsm"):
        return probe_xlsx(path)
    if suffix == ".csv":
        return probe_csv(path)
    if suffix == ".xls":
        return FileShape(
            path=str(path),
            kind="unknown",
            note="구형 .xls 형식입니다. openpyxl 로 못 읽으니 xlrd 또는 변환이 필요합니다.",
        )
    return FileShape(path=str(path), kind="unknown", note=f"처리 규칙이 없는 확장자: {suffix}")


def probe_all(paths: list[Path]) -> list[FileShape]:
    return [probe(p) for p in paths]


def to_json(shapes: list[FileShape]) -> str:
    return json.dumps([asdict(s) for s in shapes], ensure_ascii=False, indent=2)
