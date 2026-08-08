"""뱅크샐러드 '가계부 내역' 시트 파서.

실측 스키마 (2026-08-08 내보내기 기준):
  헤더 1행, 10컬럼 — 날짜 · 시간 · 타입 · 대분류 · 소분류 · 내용 · 금액 · 화폐 · 결제수단 · 메모
  타입은 {수입, 지출, 이체} 3종. 메모는 항상 비어 있음.

⚠️ 조용히 틀린 숫자를 만드는 함정 세 개 — 셋 다 그럴듯한 값이라 눈으로 안 걸린다.

1. '이체'는 내 계좌 간 이동이라 수입에도 지출에도 넣으면 안 된다.
   빼지 않으면 수입과 지출이 동시에 부풀려진다. (실측: 1년치 509건 912만원)

2. 뱅샐은 지출을 **음수**로 기록한다. 부호를 그대로 합산하면
   순저축 = 수입 - (음수) = 수입 + 지출액 이 되어 저축률이 수백 %로 튄다.

3. **타입이 '지출'인데 금액이 양수인 행은 환불·취소다.** (실측 36건 71만원)
   2번을 abs() 로 때우면 환불이 지출로 더해져 오차가 두 배로 벌어진다.

   → 해결: amount 는 **원본 부호를 그대로** 보존하고, 집계에서 부호를 뒤집는다.
     지출 순액 = -Σ(지출 행의 amount). 음수는 지출로, 양수(환불)는 자동 차감된다.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

LEDGER_SHEET = "가계부 내역"
EXPECTED_HEADER = [
    "날짜", "시간", "타입", "대분류", "소분류", "내용", "금액", "화폐", "결제수단", "메모",
]

INCOME = "수입"
EXPENSE = "지출"
TRANSFER = "이체"


class LedgerSchemaError(RuntimeError):
    pass


@dataclass(frozen=True)
class Txn:
    day: date
    kind: str  # 수입 | 지출 | 이체
    major: str  # 대분류
    minor: str  # 소분류
    desc: str  # 내용
    amount: int  # 원본 부호 그대로. 지출은 음수, 환불은 양수.
    currency: str
    method: str  # 결제수단

    @property
    def month(self) -> str:
        return self.day.strftime("%Y-%m")

    @property
    def key(self) -> tuple[str, str, str]:
        """고정지출 판정을 위한 반복 항목 식별자."""
        return (self.major, self.minor, self.desc)

    @property
    def is_refund(self) -> bool:
        """지출로 분류됐지만 금액이 양수 — 환불·취소."""
        return self.kind == EXPENSE and self.amount > 0

    @property
    def outflow(self) -> int:
        """지출 방향의 순액. 환불이면 음수가 되어 자동으로 차감된다."""
        return -self.amount if self.kind == EXPENSE else 0


def _text(v) -> str:
    return "" if v is None else str(v).strip()


def load(xlsx: Path) -> list[Txn]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        if LEDGER_SHEET not in wb.sheetnames:
            raise LedgerSchemaError(
                f"'{LEDGER_SHEET}' 시트가 없습니다. 발견된 시트: {wb.sheetnames}"
            )
        ws = wb[LEDGER_SHEET]
        rows = ws.iter_rows(values_only=True)

        header = [_text(c) for c in next(rows, [])]
        if header[: len(EXPECTED_HEADER)] != EXPECTED_HEADER:
            # 뱅샐이 컬럼을 바꾸면 조용히 틀린 값을 내놓는 대신 여기서 멈춘다.
            raise LedgerSchemaError(
                f"헤더가 예상과 다릅니다.\n  예상: {EXPECTED_HEADER}\n  실제: {header}"
            )

        txns: list[Txn] = []
        for r in rows:
            if not r or r[0] is None:
                continue
            day = r[0].date() if hasattr(r[0], "date") else r[0]
            txns.append(
                Txn(
                    day=day,
                    kind=_text(r[2]),
                    major=_text(r[3]),
                    minor=_text(r[4]),
                    desc=_text(r[5]),
                    amount=int(r[6] or 0),
                    currency=_text(r[7]),
                    method=_text(r[8]),
                )
            )
        return txns
    finally:
        wb.close()


# ── 집계 ─────────────────────────────────────────────────────────────


def months(txns: list[Txn]) -> list[str]:
    return sorted({t.month for t in txns})


def monthly_totals(txns: list[Txn]) -> dict[str, dict[str, int]]:
    """월별 수입/지출 순액. 이체는 제외하고, 환불은 지출에서 차감한다."""
    out: dict[str, dict[str, int]] = defaultdict(lambda: {INCOME: 0, EXPENSE: 0})
    for t in txns:
        if t.kind == INCOME:
            out[t.month][INCOME] += t.amount
        elif t.kind == EXPENSE:
            out[t.month][EXPENSE] += t.outflow
    return dict(out)


def refunds(txns: list[Txn]) -> list[Txn]:
    """환불·취소 건. 감사 리포트에서 따로 보여줄 가치가 있다."""
    return [t for t in txns if t.is_refund]


def split_fixed_variable(
    txns: list[Txn], *, min_months: int = 3, tolerance: float = 0.15
) -> tuple[set[tuple[str, str, str]], set[tuple[str, str, str]]]:
    """지출을 고정/변동으로 나눈다.

    고정지출 판정: 같은 (대분류·소분류·내용) 조합이 min_months 개월 이상 등장하고,
    월별 순액의 변동폭이 평균 대비 tolerance 이내인 것.

    휴리스틱이다. 구독료·월세·보험료를 잡는 것이 목적이고,
    매달 사먹는 커피처럼 금액이 들쭉날쭉한 건 변동으로 남는다.
    """
    by_key: dict[tuple[str, str, str], list[Txn]] = defaultdict(list)
    for t in txns:
        if t.kind == EXPENSE:
            by_key[t.key].append(t)

    fixed: set[tuple[str, str, str]] = set()
    variable: set[tuple[str, str, str]] = set()

    for key, group in by_key.items():
        # 같은 달에 두 번 결제되거나 환불이 섞인 경우를 위해 월별 순액으로 판단
        per_month: dict[str, int] = defaultdict(int)
        for t in group:
            per_month[t.month] += t.outflow
        if len(per_month) < min_months:
            variable.add(key)
            continue
        amounts = list(per_month.values())
        avg = sum(amounts) / len(amounts)
        if avg <= 0:  # 순액이 0 이하 = 전액 환불된 항목
            variable.add(key)
            continue
        spread = (max(amounts) - min(amounts)) / avg
        (fixed if spread <= tolerance else variable).add(key)

    return fixed, variable


def category_totals(txns: list[Txn], kind: str = EXPENSE) -> dict[str, int]:
    """대분류별 순액. 지출은 환불 차감 후, 수입은 그대로."""
    out: dict[str, int] = defaultdict(int)
    for t in txns:
        if t.kind != kind:
            continue
        out[t.major] += t.outflow if kind == EXPENSE else t.amount
    return dict(sorted(out.items(), key=lambda kv: kv[1], reverse=True))
