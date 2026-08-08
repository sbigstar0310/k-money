"""뱅크샐러드 '가계부 내역' 시트 파서.

실측 스키마 (2026-08-08 내보내기 기준):
  헤더 1행, 10컬럼 — 날짜 · 시간 · 타입 · 대분류 · 소분류 · 내용 · 금액 · 화폐 · 결제수단 · 메모
  타입은 {수입, 지출, 이체} 3종. 메모는 항상 비어 있음.

⚠️ 함정 두 개 — 둘 다 조용히 틀린 숫자를 만든다.

1. '이체'는 내 계좌 간 이동이라 수입에도 지출에도 넣으면 안 된다.
   빼지 않으면 수입과 지출이 동시에 부풀려진다. (실측: 1년치 509건 912만원)

2. 뱅샐은 지출 금액을 **음수**로 기록한다. 부호를 그대로 합산하면
   순저축 = 수입 - (음수) = 수입 + 지출액 이 되어 저축률이 수백 %로 튄다.
   여기서는 amount 를 항상 양수 크기로 정규화하고, 방향은 kind 가 갖는다.
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

LEDGER_SHEET = "가계부 내역"
EXPECTED_HEADER = [
    "날짜", "시간", "타입", "대분류", "소분류", "내용", "금액", "화폐", "결제수단", "메모",
]

INCOME = "수입"
EXPENSE = "지출"
TRANSFER = "이체"


class LedgerSchemaError(RuntimeError):
    pass


@dataclass(frozen=True)
class Txn:
    day: date
    kind: str  # 수입 | 지출 | 이체
    major: str  # 대분류
    minor: str  # 소분류
    desc: str  # 내용
    amount: int  # 항상 양수 크기. 방향은 kind 가 갖는다 (원본은 지출이 음수)
    currency: str
    method: str  # 결제수단

    @property
    def month(self) -> str:
        return self.day.strftime("%Y-%m")

    @property
    def key(self) -> tuple[str, str, str]:
        """고정지출 판정을 위한 반복 항목 식별자."""
        return (self.major, self.minor, self.desc)


def _text(v) -> str:
    return "" if v is None else str(v).strip()


def load(xlsx: Path) -> list[Txn]:
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    try:
        if LEDGER_SHEET not in wb.sheetnames:
            raise LedgerSchemaError(
                f"'{LEDGER_SHEET}' 시트가 없습니다. 발견된 시트: {wb.sheetnames}"
            )
        ws = wb[LEDGER_SHEET]
        rows = ws.iter_rows(values_only=True)

        header = [_text(c) for c in next(rows, [])]
        if header[: len(EXPECTED_HEADER)] != EXPECTED_HEADER:
            # 뱅샐이 컬럼을 바꾸면 조용히 틀린 값을 내놓는 대신 여기서 멈춘다.
            raise LedgerSchemaError(
                f"헤더가 예상과 다릅니다.\n  예상: {EXPECTED_HEADER}\n  실제: {header}"
            )

        txns: list[Txn] = []
        for r in rows:
            if not r or r[0] is None:
                continue
            day = r[0].date() if hasattr(r[0], "date") else r[0]
            txns.append(
                Txn(
                    day=day,
                    kind=_text(r[2]),
                    major=_text(r[3]),
                    minor=_text(r[4]),
                    desc=_text(r[5]),
                    amount=abs(int(r[6] or 0)),
                    currency=_text(r[7]),
                    method=_text(r[8]),
                )
            )
        return txns
    finally:
        wb.close()


# ── 집계 ─────────────────────────────────────────────────────────────


def months(txns: list[Txn]) -> list[str]:
    return sorted({t.month for t in txns})


def monthly_totals(txns: list[Txn]) -> dict[str, dict[str, int]]:
    """월별 수입/지출 합계. 이체는 제외한다."""
    out: dict[str, dict[str, int]] = defaultdict(lambda: {INCOME: 0, EXPENSE: 0})
    for t in txns:
        if t.kind == TRANSFER:
            continue
        out[t.month][t.kind] += t.amount
    return dict(out)


def split_fixed_variable(
    txns: list[Txn], *, min_months: int = 3, tolerance: float = 0.15
) -> tuple[set[tuple[str, str, str]], set[tuple[str, str, str]]]:
    """지출을 고정/변동으로 나눈다.

    고정지출 판정: 같은 (대분류·소분류·내용) 조합이 min_months 개월 이상 등장하고,
    금액의 변동폭이 평균 대비 tolerance 이내인 것.

    휴리스틱이다. 구독료·월세·보험료를 잡는 것이 목적이고,
    매달 사먹는 커피처럼 금액이 들쭉날쭉한 건 변동으로 남는다.
    """
    by_key: dict[tuple[str, str, str], list[Txn]] = defaultdict(list)
    for t in txns:
        if t.kind == EXPENSE:
            by_key[t.key].append(t)

    fixed: set[tuple[str, str, str]] = set()
    variable: set[tuple[str, str, str]] = set()

    for key, group in by_key.items():
        distinct_months = {t.month for t in group}
        if len(distinct_months) < min_months:
            variable.add(key)
            continue
        # 월별 합계 기준으로 안정성 판단 (같은 달에 두 번 결제되는 경우 대비)
        per_month: dict[str, int] = defaultdict(int)
        for t in group:
            per_month[t.month] += t.amount
        amounts = list(per_month.values())
        avg = sum(amounts) / len(amounts)
        if avg <= 0:
            variable.add(key)
            continue
        spread = (max(amounts) - min(amounts)) / avg
        (fixed if spread <= tolerance else variable).add(key)

    return fixed, variable


def category_totals(txns: list[Txn], kind: str = EXPENSE) -> dict[str, int]:
    out: dict[str, int] = defaultdict(int)
    for t in txns:
        if t.kind == kind:
            out[t.major] += t.amount
    return dict(sorted(out.items(), key=lambda kv: kv[1], reverse=True))
