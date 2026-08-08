"""개발 도구 — xlsx 의 시트를 그대로 JSON 행 배열로 뽑는다.

프로덕션에는 안 들어간다. Apps Script 는 SpreadsheetApp 의 getValues() 로
같은 모양의 행을 얻으므로, 이 파일은 **로컬에서 core/ 를 검증하기 위한
픽스처 생성기**일 뿐이다.

    pip install openpyxl
    python3 scripts/dump-rows.py data/extracted/*.xlsx fixtures/real.json

⚠️ 이 저장소에 남은 **유일한 파이썬**이다. 파이프라인은 전부 Apps Script 로
   갔고 src/ 는 지웠다. 이 파일만 남긴 이유는 xlsx 를 읽는 데 openpyxl 만한
   게 없어서다 — 그리고 실데이터 픽스처가 없으면 smoke-real 테스트가 못 돈다.
   (합성 데이터가 스펙이고, 실데이터는 있으면 돌리는 보조 검증이다.)

날짜는 getValues() 가 Date 를 주는 것과 맞추기 위해 ISO 문자열로 내보낸다
(core/parse.js 의 day() 가 Date 와 문자열을 둘 다 받는다).
"""

from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def cellify(v):
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    if v is None or isinstance(v, (str, int, float, bool)):
        return v
    return str(v)


def main(argv: list[str]) -> int:
    if len(argv) != 2:
        print("사용법: dump-rows.py <입력.xlsx> <출력.json>", file=sys.stderr)
        return 2

    src, dest = Path(argv[0]), Path(argv[1])
    wb = load_workbook(src, read_only=True, data_only=True)
    try:
        sheets = {
            name: [[cellify(c) for c in row]
                   for row in wb[name].iter_rows(values_only=True)]
            for name in wb.sheetnames
        }
    finally:
        wb.close()

    dest.parent.mkdir(parents=True, exist_ok=True)
    dest.write_text(json.dumps(sheets, ensure_ascii=False), encoding="utf-8")

    for name, rows in sheets.items():
        print(f"  {name}: {len(rows)}행")
    print(f"→ {dest} ({dest.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv[1:]))
